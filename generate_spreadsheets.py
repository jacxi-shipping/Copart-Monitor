"""
Copart Spreadsheet Generator
Reads state.json, watchlist.json, and watchlist_archive.json and produces:
  - copart_lots.xlsx      → All seen lots with full details
  - copart_auctions.xlsx  → Bid snapshots, history, and final bid analysis
"""
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")  # Dark blue
SUBHEAD_FILL  = PatternFill("solid", start_color="2E75B6")  # Mid blue
ALT_FILL      = PatternFill("solid", start_color="EBF3FB")  # Light blue row
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
GREEN_FILL    = PatternFill("solid", start_color="E2EFDA")
RED_FILL      = PatternFill("solid", start_color="FCE4D6")
YELLOW_FILL   = PatternFill("solid", start_color="FFF2CC")

HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
SMALL_FONT    = Font(name="Arial", size=9)

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _monitor_label() -> str:
    """Build a short label from COPART_MAKES / COPART_MODELS env vars."""
    makes = [m.strip() for m in os.environ.get("COPART_MAKES", "").split(",") if m.strip()]
    models = [m.strip() for m in os.environ.get("COPART_MODELS", "").split(",") if m.strip()]
    parts = makes[:1] + models[:2]   # e.g. ["Toyota", "RAV4", "RAV4 HYBRID"]
    if parts:
        return " / ".join(parts)
    return "Vehicle"


def _header(ws, col, row, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row, col, value, fill=None, font=None, align=None, num_fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or (Font(name="Arial", size=10, bold=bold) if bold else NORMAL_FONT)
    c.fill = fill or (ALT_FILL if row % 2 == 0 else WHITE_FILL)
    c.alignment = align or LEFT
    c.border = THIN_BORDER
    if num_fmt:
        c.number_format = num_fmt
    return c


def _ts_to_str(ts_ms) -> str:
    if not ts_ms:
        return ""
    try:
        ts = int(ts_ms)
        if ts > 1_000_000_000_000:
            ts /= 1000
        return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        return str(ts_ms)


def _iso_to_str(iso) -> str:
    if not iso:
        return ""
    try:
        return iso[:19].replace("T", " ") + " UTC"
    except Exception:
        return str(iso)


# ---------------------------------------------------------------------------
# Sheet 1 — All Seen Lots
# ---------------------------------------------------------------------------

LOT_COLUMNS = [
    ("Lot #",           12, "lot_number"),
    ("Title",           32, "title"),
    ("Year",             7, "year"),
    ("Make",            12, "make"),
    ("Model",           14, "model"),
    ("Trim",            14, "trim"),
    ("Drive Status",    18, "drive_status"),
    ("Has Keys",        10, "has_keys"),
    ("Damage",          20, "damage"),
    ("Secondary Dmg",   18, "secondary_damage"),
    ("Odometer",        12, "odometer"),
    ("Est. Value ($)",  14, "estimate"),
    ("Current Bid ($)", 14, "current_bid"),
    ("NLR / Broker",    14, "is_nlr"),
    ("Location",        22, "location"),
    ("Sale Date",       22, "sale_date"),
    ("Title Type",      18, "title_type"),
    ("VIN",             20, "vin"),
    ("Engine",          14, "engine"),
    ("First Seen",      22, "first_seen"),
    ("URL",             50, "url"),
]


def build_lots_sheet(wb, lot_details: dict):
    ws = wb.create_sheet("All Seen Lots")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # Headers
    for col_idx, (label, width, _) in enumerate(LOT_COLUMNS, 1):
        _header(ws, col_idx, 1, label, width)

    # Data rows
    for row_idx, (lot_num, lot) in enumerate(sorted(lot_details.items()), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL

        # Drive status colouring
        drive = (lot.get("drive_status") or "").upper()
        if "RUNS AND DRIVES" in drive or "RUN AND DRIVE" in drive:
            drive_fill = GREEN_FILL
        elif "STATIONARY" in drive:
            drive_fill = RED_FILL
        else:
            drive_fill = fill

        for col_idx, (_, _, field) in enumerate(LOT_COLUMNS, 1):
            raw = lot.get(field)

            # Format specific fields
            if field == "is_nlr":
                value = "✅ NLR" if raw else "🔑 Broker"
            elif field == "has_keys":
                value = "Yes" if raw is True else ("No" if raw is False else "")
            elif field == "sale_date":
                value = _ts_to_str(raw)
            elif field == "first_seen":
                value = _iso_to_str(raw)
            elif field in ("odometer", "estimate", "current_bid"):
                try:
                    value = int(raw) if raw not in (None, "") else ""
                except (ValueError, TypeError):
                    value = raw
            else:
                value = raw if raw is not None else ""

            cell_fill = drive_fill if field == "drive_status" else fill
            num_fmt = "#,##0" if field in ("odometer", "estimate", "current_bid") else None
            _cell(ws, row_idx, col_idx, value, fill=cell_fill, num_fmt=num_fmt)

    # Summary row at top after freeze
    total = len(lot_details)
    nlr_count = sum(1 for l in lot_details.values() if l.get("is_nlr"))
    runs_count = sum(1 for l in lot_details.values()
                     if "RUNS AND DRIVES" in (l.get("drive_status") or "").upper()
                     or "RUN AND DRIVE" in (l.get("drive_status") or "").upper())

    # Add summary info at the top using a title row
    ws.insert_rows(1)
    ws.row_dimensions[1].height = 18
    title_cell = ws.cell(row=1, column=1, value=f"Copart {_monitor_label()} Monitor — All Seen Lots")
    title_cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="1A3A5C")
    title_cell.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(LOT_COLUMNS))

    # Re-freeze after insert
    ws.freeze_panes = "A3"

    # Summary below headers
    summary_row = len(lot_details) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = BOLD_FONT
    ws.cell(row=summary_row, column=2, value=f"Total Lots: {total}").font = NORMAL_FONT
    ws.cell(row=summary_row, column=3, value=f"NLR: {nlr_count}").font = Font(name="Arial", size=10, color="00703C")
    ws.cell(row=summary_row, column=4, value=f"Broker: {total - nlr_count}").font = NORMAL_FONT
    ws.cell(row=summary_row, column=5, value=f"Runs & Drives: {runs_count}").font = Font(name="Arial", size=10, color="00703C")

    return ws


# ---------------------------------------------------------------------------
# Sheet 2 — Active Watchlist (current bids)
# ---------------------------------------------------------------------------

def build_watchlist_sheet(wb, watchlist: dict, title="Active Watchlist"):
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A3"

    headers = [
        ("Lot #", 12), ("Title", 30), ("Year", 7), ("Drive Status", 18),
        ("Target ($)", 12), ("Last Bid ($)", 12), ("# Snapshots", 12),
        ("Min Bid ($)", 12), ("Max Bid ($)", 12), ("Bid Trend", 14),
        ("NLR/Broker", 12), ("Sale Date", 22), ("Added At", 22), ("URL", 45),
    ]

    # Title row
    ws.row_dimensions[1].height = 18
    tc = ws.cell(row=1, column=1, value=f"Copart {_monitor_label()} Tracker — {title}")
    tc.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color="1A3A5C")
    tc.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    ws.row_dimensions[2].height = 22
    for col_idx, (label, width) in enumerate(headers, 1):
        _header(ws, col_idx, 2, label, width)

    for row_idx, (ln, lot) in enumerate(watchlist.items(), 3):
        history = lot.get("bid_history") or []
        bids = [s["bid"] for s in history if s.get("bid") is not None]

        target = lot.get("target_price", 0) or 0
        last_bid = lot.get("last_bid") or (bids[-1] if bids else None)
        min_bid = min(bids) if bids else None
        max_bid = max(bids) if bids else None
        num_snaps = len(bids)

        # Bid trend
        trend = ""
        if len(bids) >= 2:
            diff = bids[-1] - bids[-2]
            trend = f"↑ +${diff:,.0f}" if diff > 0 else (f"→ No change" if diff == 0 else f"↓ ${diff:,.0f}")

        # Row fill — green if last bid under target, red if over
        if last_bid is not None and target:
            row_fill = GREEN_FILL if last_bid <= target else RED_FILL
        else:
            row_fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL

        drive = (lot.get("drive_status") or "").title()

        vals = [
            ln, lot.get("title", ""), lot.get("year", ""),
            drive,
            target, last_bid, num_snaps,
            min_bid, max_bid, trend,
            "✅ NLR" if lot.get("is_nlr") else "🔑 Broker",
            _ts_to_str(lot.get("sale_date")),
            _iso_to_str(lot.get("added_at")),
            lot.get("url", ""),
        ]
        num_fields = {4, 5, 7, 8}  # indices (1-based) that are currency
        for col_idx, val in enumerate(vals, 1):
            num_fmt = "#,##0" if col_idx in num_fields else None
            _cell(ws, row_idx, col_idx, val, fill=row_fill, num_fmt=num_fmt)

    return ws


# ---------------------------------------------------------------------------
# Sheet 3 — Bid Snapshot History (one row per snapshot)
# ---------------------------------------------------------------------------

def build_bid_history_sheet(wb, all_lots: dict):
    ws = wb.create_sheet("Bid Snapshots")
    ws.freeze_panes = "A3"

    headers = [
        ("Lot #", 12), ("Title", 30), ("Year", 7),
        ("Snapshot #", 12), ("Timestamp (UTC)", 22), ("Bid ($)", 12),
        ("Target ($)", 12), ("Under/Over Target", 18),
        ("Status", 14), ("Drive Status", 18), ("NLR/Broker", 12),
    ]

    tc = ws.cell(row=1, column=1, value=f"Copart {_monitor_label()} Tracker — Bid Snapshot History")
    tc.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color="1A3A5C")
    tc.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    for col_idx, (label, width) in enumerate(headers, 1):
        _header(ws, col_idx, 2, label, width)

    row_idx = 3
    for ln, lot in all_lots.items():
        history = lot.get("bid_history") or []
        target = lot.get("target_price", 0) or 0
        drive = (lot.get("drive_status") or "").title()
        status = lot.get("auction_result") or ("Closed" if lot.get("final_bid") is not None else "Active")

        for snap_num, snap in enumerate(history, 1):
            bid = snap.get("bid") or 0
            under_over = target - bid
            under_over_str = f"+${under_over:,.0f} under" if under_over >= 0 else f"-${-under_over:,.0f} over"
            fill = GREEN_FILL if under_over >= 0 else RED_FILL

            vals = [
                ln,
                lot.get("title", ""),
                lot.get("year", ""),
                snap_num,
                _iso_to_str(snap.get("timestamp")),
                bid,
                target,
                under_over_str,
                status,
                drive,
                "✅ NLR" if lot.get("is_nlr") else "🔑 Broker",
            ]
            for col_idx, val in enumerate(vals, 1):
                num_fmt = "#,##0" if col_idx in (6, 7) else None
                row_fill = fill if col_idx in (6, 8) else (ALT_FILL if row_idx % 2 == 0 else WHITE_FILL)
                _cell(ws, row_idx, col_idx, val, fill=row_fill, num_fmt=num_fmt)
            row_idx += 1

    return ws


# ---------------------------------------------------------------------------
# Sheet 4 — Final Bid Analysis (closed auctions)
# ---------------------------------------------------------------------------

def build_final_bid_sheet(wb, archive: dict):
    ws = wb.create_sheet("Final Bid Analysis")
    ws.freeze_panes = "A3"

    headers = [
        ("Lot #", 12), ("Title", 30), ("Year", 7),
        ("Drive Status", 18), ("Damage", 20), ("Odometer", 12),
        ("NLR/Broker", 12), ("Est. Value ($)", 14),
        ("Your Target ($)", 14), ("Final Bid ($)", 14),
        ("Won/Lost Budget", 16), ("# Bid Snapshots", 14),
        ("Opening Bid ($)", 14), ("Bid Increase", 14),
        ("Auction Result", 14), ("Closed At", 22), ("URL", 45),
    ]

    tc = ws.cell(row=1, column=1, value=f"Copart {_monitor_label()} Tracker — Final Bid Analysis (Closed Auctions)")
    tc.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color="1A3A5C")
    tc.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    for col_idx, (label, width) in enumerate(headers, 1):
        _header(ws, col_idx, 2, label, width)

    row_idx = 3
    for ln, lot in archive.items():
        history = lot.get("bid_history") or []
        bids = [s["bid"] for s in history if s.get("bid") is not None]

        final_bid = lot.get("final_bid") or (bids[-1] if bids else None)
        opening_bid = bids[0] if bids else None
        target = lot.get("target_price", 0) or 0
        estimate = lot.get("estimate")
        drive = (lot.get("drive_status") or "").title()
        result = lot.get("auction_result", "CLOSED")

        # Win/loss assessment
        if final_bid is not None and target:
            diff = target - final_bid
            won_lost = f"+${diff:,.0f} under" if diff >= 0 else f"-${-diff:,.0f} over"
            row_fill = GREEN_FILL if diff >= 0 else RED_FILL
        else:
            won_lost = ""
            row_fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL

        # Bid increase from open to close
        bid_increase = ""
        if opening_bid is not None and final_bid is not None:
            bid_increase = f"${final_bid - opening_bid:,.0f}"

        vals = [
            ln,
            lot.get("title", ""),
            lot.get("year", ""),
            drive,
            lot.get("damage", ""),
            lot.get("odometer"),
            "✅ NLR" if lot.get("is_nlr") else "🔑 Broker",
            estimate,
            target,
            final_bid,
            won_lost,
            len(bids),
            opening_bid,
            bid_increase,
            result,
            _iso_to_str(lot.get("closed_at")),
            lot.get("url", ""),
        ]
        currency_cols = {8, 9, 10, 13}
        for col_idx, val in enumerate(vals, 1):
            num_fmt = "#,##0" if col_idx in currency_cols else None
            cell_fill = row_fill if col_idx in (10, 11) else (ALT_FILL if row_idx % 2 == 0 else WHITE_FILL)
            _cell(ws, row_idx, col_idx, val, fill=cell_fill, num_fmt=num_fmt)

        row_idx += 1

    # Summary stats at bottom
    if archive:
        final_bids = [lot.get("final_bid") for lot in archive.values() if lot.get("final_bid") is not None]
        if final_bids:
            avg_final = sum(final_bids) / len(final_bids)
            summary_row = row_idx + 1
            ws.cell(row=summary_row, column=1, value="ANALYSIS").font = BOLD_FONT
            ws.cell(row=summary_row, column=2, value=f"Avg Final Bid: ${avg_final:,.0f}").font = NORMAL_FONT
            ws.cell(row=summary_row, column=3, value=f"Min: ${min(final_bids):,.0f}").font = NORMAL_FONT
            ws.cell(row=summary_row, column=4, value=f"Max: ${max(final_bids):,.0f}").font = NORMAL_FONT
            ws.cell(row=summary_row, column=5, value=f"Samples: {len(final_bids)}").font = NORMAL_FONT

    return ws


# ---------------------------------------------------------------------------
# Main entry points
# ---------------------------------------------------------------------------

def generate_lots_spreadsheet(state_file="state.json", output="copart_lots.xlsx"):
    state = {}
    p = Path(state_file)
    if p.exists():
        try:
            state = json.loads(p.read_text())
        except Exception as e:
            print(f"Error reading {state_file}: {e}")
            return

    lot_details = state.get("lot_details", {})
    if not lot_details:
        print("No lot details found in state.json — run the monitor first")
        return

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    build_lots_sheet(wb, lot_details)
    wb.save(output)
    print(f"✅ Lots spreadsheet saved: {output} ({len(lot_details)} lots)")


def generate_auction_spreadsheet(watchlist_file="watchlist.json",
                                  archive_file="watchlist_archive.json",
                                  output="copart_auctions.xlsx"):
    watchlist = {}
    archive = {}

    if Path(watchlist_file).exists():
        try:
            watchlist = json.loads(Path(watchlist_file).read_text())
        except Exception as e:
            print(f"Error reading {watchlist_file}: {e}")

    if Path(archive_file).exists():
        try:
            archive = json.loads(Path(archive_file).read_text())
        except Exception as e:
            print(f"Error reading {archive_file}: {e}")

    all_lots = {**archive, **watchlist}  # archive first, then active on top

    wb = Workbook()
    wb.remove(wb.active)
    build_watchlist_sheet(wb, watchlist, title="Active Watchlist")
    build_watchlist_sheet(wb, archive, title="Closed Auctions")
    build_bid_history_sheet(wb, all_lots)
    build_final_bid_sheet(wb, archive)
    wb.save(output)
    print(f"✅ Auction spreadsheet saved: {output} "
          f"({len(watchlist)} active, {len(archive)} closed)")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--lots", action="store_true", help="Generate lots spreadsheet")
    parser.add_argument("--auctions", action="store_true", help="Generate auctions spreadsheet")
    parser.add_argument("--all", action="store_true", help="Generate both spreadsheets")
    args = parser.parse_args()

    if args.lots or args.all:
        generate_lots_spreadsheet()
    if args.auctions or args.all:
        generate_auction_spreadsheet()
    if not any([args.lots, args.auctions, args.all]):
        # Default: generate both
        generate_lots_spreadsheet()
        generate_auction_spreadsheet()
